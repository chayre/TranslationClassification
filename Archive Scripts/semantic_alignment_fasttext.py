"""
semantic_alignment_fasttext.py
===============================
Implements the semantic alignment algorithm from Liu et al. (2025)
using fastText embeddings and a Polish-English dictionary.

The method works regardless of whether the vector spaces are aligned,
because it compares similarity PROFILES rather than direct vectors.

Usage:
  python semantic_alignment_fasttext.py
  python semantic_alignment_fasttext.py --top 10
"""

import argparse
import csv
import os
import re
import sys
from collections import defaultdict

import numpy as np
from scipy.stats import pearsonr

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────

BASE_DIR = r"C:\Users\CAyre\Documents\Coding\TranslationClassification\TranslationClassification"
DATA_DIR = os.path.join(BASE_DIR, "data")
RESULTS_DIR = os.path.join(BASE_DIR, "results")

POLISH_CDI = os.path.join(DATA_DIR, "cdi_pl_diminutives.csv")
ENGLISH_CDI = os.path.join(DATA_DIR, "american_english_itemdata.csv")
DICTIONARY = os.path.join(DATA_DIR, "pl_en_dict.csv")

FASTTEXT_PL = os.path.join(DATA_DIR, "cc.pl.300.vec")
FASTTEXT_EN = os.path.join(DATA_DIR, "cc.en.300.vec")

OUTPUT_CSV = os.path.join(DATA_DIR, "semantic_alignment_scores.csv")
OUTPUT_EN_FOR_PL = os.path.join(RESULTS_DIR, "top10_english_for_polish_sa.xlsx")
OUTPUT_PL_FOR_EN = os.path.join(RESULTS_DIR, "top10_polish_for_english_sa.xlsx")

# ── Excel styles ──────────────────────────────────────────────────────────────

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BODY = Font(name="Arial", size=10)


def get_fill(score):
    if score >= 0.75:
        return FILL_GREEN
    elif score >= 0.60:
        return FILL_YELLOW
    else:
        return FILL_RED


# ── helpers ───────────────────────────────────────────────────────────────────

def strip_parens(text):
    return re.sub(r"\s*\(.*?\)", "", text).strip()


def load_cdi(path):
    """Load CDI words from CSV. Expects 'item_definition' and 'category' columns."""
    seen, items = set(), []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            word = strip_parens(row["item_definition"])
            cat = row.get("category", "")
            if word not in seen:
                seen.add(word)
                items.append((word, cat))
    return items


def load_dictionary(path):
    """
    Load the shared associate dictionary.
    Expects CSV with 'english' and 'polish' columns.
    Handles one-to-many mappings (one English word -> multiple Polish words).
    """
    mapping = defaultdict(list)
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            en_word = strip_parens(row["english"]).lower()
            pl_word = strip_parens(row["polish"]).lower()

            mapping[en_word].append(pl_word)
            if row.get("polish_alternatives"):
                for alt in row["polish_alternatives"].split("|"):
                    alt = strip_parens(alt).lower()
                    if alt and alt != pl_word:
                        mapping[en_word].append(alt)

    associates = []
    for en_word, pl_words in mapping.items():
        pl_words = list(set(pl_words))
        associates.append({"en": en_word, "pl_list": pl_words})

    return associates


# ── fastText loading ─────────────────────────────────────────────────────────

def load_fasttext_vec(path, target_words):
    """
    Load a fastText .vec file, keeping only vectors for target_words.
    Returns dict: {word: normalized_vector}
    """
    print(f"  Loading {os.path.basename(path)}...")

    word_set = set(w.lower() for w in target_words)
    word_vecs = {}

    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        first_line = f.readline().strip()
        parts = first_line.split()
        if len(parts) == 2:
            try:
                vocab_size, dim = map(int, parts)
                print(f"    Header: {vocab_size} words, dim={dim}")
            except ValueError:
                f.seek(0)
        else:
            f.seek(0)

        for line in f:
            if not line.strip():
                continue
            parts = line.rstrip().split(' ')
            word = parts[0].lower()
            if word in word_set and word not in word_vecs:
                try:
                    vec = np.array([float(x) for x in parts[1:]], dtype=np.float32)
                    norm = np.linalg.norm(vec)
                    if norm > 0:
                        word_vecs[word] = vec / norm
                except ValueError:
                    continue

    missing = len(word_set) - len(word_vecs)
    print(f"    Found {len(word_vecs)}/{len(word_set)} words"
          + (f" ({missing} missing)" if missing > 0 else ""))

    return word_vecs


# ── core algorithm ───────────────────────────────────────────────────────────

def build_associate_vectors(associates, pl_vecs, en_vecs):
    """
    For each shared associate, get its English vector and the average
    of its Polish equivalent vectors (Formula 1 from Liu et al., 2025).
    """
    en_vec_list = []
    pl_vec_list = []

    for assoc in associates:
        en_word = assoc["en"].lower()
        pl_words = [w.lower() for w in assoc["pl_list"]]

        en_vec = en_vecs.get(en_word)
        if en_vec is None:
            continue

        pl_vecs_for_word = []
        for pw in pl_words:
            if pw in pl_vecs:
                pl_vecs_for_word.append(pl_vecs[pw])

        if not pl_vecs_for_word:
            continue

        pl_vec = np.mean(pl_vecs_for_word, axis=0)
        pl_vec = pl_vec / (np.linalg.norm(pl_vec) + 1e-10)

        en_vec_list.append(en_vec)
        pl_vec_list.append(pl_vec)

    print(f"  Valid shared associates: {len(en_vec_list)}/{len(associates)}")
    return np.array(en_vec_list), np.array(pl_vec_list)


def semantic_alignment_score(pl_word, en_word, pl_vecs, en_vecs,
                              en_assoc_vecs, pl_assoc_vecs):
    """
    Calculate Pearson R_c between two words' similarity profiles.
    This is the core metric from Liu et al. (2025).
    """
    pl_vec = pl_vecs.get(pl_word.lower())
    en_vec = en_vecs.get(en_word.lower())

    if pl_vec is None or en_vec is None:
        return None

    sims_pl = pl_assoc_vecs @ pl_vec
    sims_en = en_assoc_vecs @ en_vec

    if len(sims_pl) < 3:
        return None

    r_c, _ = pearsonr(sims_en, sims_pl)
    return r_c


# ── main pipeline (updated end section) ─────────────────────────────────────

def run_pipeline(top_n=10):
    # 1. Load data
    print("=" * 60)
    print("STEP 1: Loading CDI data and dictionary")
    print("=" * 60)

    pl_items = load_cdi(POLISH_CDI)
    en_items = load_cdi(ENGLISH_CDI)
    associates = load_dictionary(DICTIONARY)

    pl_words = [w for w, _ in pl_items]
    en_words = [w for w, _ in en_items]

    print(f"  Polish CDI words: {len(pl_words)}")
    print(f"  English CDI words: {len(en_words)}")
    print(f"  Dictionary entries: {len(associates)}")
    print(f"  Total pairs to score: {len(pl_words) * len(en_words):,}\n")

    # 2. Load fastText vectors
    print("=" * 60)
    print("STEP 2: Loading fastText vectors")
    print("=" * 60)

    dict_en_words = [a["en"] for a in associates]
    dict_pl_words = []
    for a in associates:
        dict_pl_words.extend(a["pl_list"])

    all_pl = list(set(pl_words + dict_pl_words))
    all_en = list(set(en_words + dict_en_words))

    print(f"  Polish words to load: {len(all_pl)}")
    print(f"  English words to load: {len(all_en)}")

    pl_vecs = load_fasttext_vec(FASTTEXT_PL, all_pl)
    en_vecs = load_fasttext_vec(FASTTEXT_EN, all_en)

    # ── NEW: Report missing words ─────────────────────────────────────────
    pl_missing = [w for w in pl_words if w.lower() not in pl_vecs]
    en_missing = [w for w in en_words if w.lower() not in en_vecs]
    dict_pl_missing = [w for w in dict_pl_words if w.lower() not in pl_vecs]
    dict_en_missing = [w for w in dict_en_words if w.lower() not in en_vecs]

    print(f"\n  ── Missing from Polish fastText ──")
    print(f"  CDI words missing: {len(pl_missing)}/{len(pl_words)}")
    if pl_missing:
        print(f"  Examples: {pl_missing[:15]}")
    print(f"  Dictionary words missing: {len(dict_pl_missing)}/{len(dict_pl_words)}")
    if dict_pl_missing:
        print(f"  Examples: {dict_pl_missing[:10]}")

    print(f"\n  ── Missing from English fastText ──")
    print(f"  CDI words missing: {len(en_missing)}/{len(en_words)}")
    if en_missing:
        print(f"  Examples: {en_missing[:15]}")
    print(f"  Dictionary words missing: {len(dict_en_missing)}/{len(dict_en_words)}")
    if dict_en_missing:
        print(f"  Examples: {dict_en_missing[:10]}")

    # Save missing words to files
    with open(os.path.join(DATA_DIR, "missing_words_pl.txt"), "w", encoding="utf-8") as f:
        f.write("CDI words missing from Polish fastText:\n")
        for w in pl_missing:
            f.write(f"  {w}\n")
        f.write(f"\nDictionary words missing from Polish fastText:\n")
        for w in sorted(set(dict_pl_missing)):
            f.write(f"  {w}\n")

    with open(os.path.join(DATA_DIR, "missing_words_en.txt"), "w", encoding="utf-8") as f:
        f.write("CDI words missing from English fastText:\n")
        for w in en_missing:
            f.write(f"  {w}\n")
        f.write(f"\nDictionary words missing from English fastText:\n")
        for w in sorted(set(dict_en_missing)):
            f.write(f"  {w}\n")

    print(f"\n  Full lists saved to data/missing_words_pl.txt and data/missing_words_en.txt")

    # 3. Build associate vectors
    print("\n" + "=" * 60)
    print("STEP 3: Building shared associate vectors")
    print("=" * 60)

    en_assoc_vecs, pl_assoc_vecs = build_associate_vectors(
        associates, pl_vecs, en_vecs
    )

    if len(en_assoc_vecs) < 3:
        print("ERROR: Too few valid shared associates. Need at least 3.")
        return

    # 4. Compute semantic alignment scores
    print("\n" + "=" * 60)
    print("STEP 4: Computing semantic alignment scores (Pearson R_c)")
    print("=" * 60)

    rows = []
    total = len(pl_words) * len(en_words)
    count = 0
    scored = 0

    for pl_w, pl_cat in pl_items:
        for en_w, en_cat in en_items:
            r_c = semantic_alignment_score(
                pl_w, en_w, pl_vecs, en_vecs,
                en_assoc_vecs, pl_assoc_vecs
            )
            if r_c is not None and not np.isnan(r_c):
                rows.append({
                    "polish_word": pl_w,
                    "english_word": en_w,
                    "polish_category": pl_cat,
                    "english_category": en_cat,
                    "R_c": round(r_c, 6),
                })
                scored += 1

            count += 1
            if count % 5000 == 0:
                print(f"  Processed {count}/{total} pairs...")

    print(f"  Scored: {scored}, Skipped: {total - scored}")

    # 5. Save CSV
    print(f"\nWriting {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "polish_word", "english_word", "polish_category",
            "english_category", "R_c"
        ])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    # 6. Excel output
    print("\n" + "=" * 60)
    print("STEP 5: Building Excel output")
    print("=" * 60)

    # Build lookup for whether a word is in fastText
    pl_in_ft = {w: (w.lower() in pl_vecs) for w in pl_words}
    en_in_ft = {w: (w.lower() in en_vecs) for w in en_words}

    pl_groups = defaultdict(list)
    en_groups = defaultdict(list)

    for r in rows:
        pl_groups[r["polish_word"]].append({
            "match_word": r["english_word"],
            "match_cat": r["english_category"],
            "index_cat": r["polish_category"],
            "score": r["R_c"],
            "match_in_ft": en_in_ft.get(r["english_word"], False),
        })
        en_groups[r["english_word"]].append({
            "match_word": r["polish_word"],
            "match_cat": r["polish_category"],
            "index_cat": r["english_category"],
            "score": r["R_c"],
            "match_in_ft": pl_in_ft.get(r["polish_word"], False),
        })

    write_excel(pl_groups, top_n, OUTPUT_EN_FOR_PL,
                "polish_word", "polish_category",
                "english_word", "english_category",
                index_in_ft=pl_in_ft)

    write_excel(en_groups, top_n, OUTPUT_PL_FOR_EN,
                "english_word", "english_category",
                "polish_word", "polish_category",
                index_in_ft=en_in_ft)

    # 7. Show top pairs
    rows_sorted = sorted(rows, key=lambda r: r["R_c"], reverse=True)
    print(f"\nTop 25 most semantically aligned pairs:")
    print(f"  {'Polish':<25} {'English':<25} {'R_c':>10}")
    print("  " + "-" * 65)
    for r in rows_sorted[:25]:
        print(f"  {r['polish_word']:<25} {r['english_word']:<25} {r['R_c']:>10.4f}")

    print("\nAll done!")
    print(f"  {OUTPUT_CSV}")
    print(f"  {OUTPUT_EN_FOR_PL}")
    print(f"  {OUTPUT_PL_FOR_EN}")


def write_excel(groups, top_n, output_path, index_col, index_cat_col,
                match_col, match_cat_col, index_in_ft=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Top matches"

    headers = [index_col, index_cat_col, "rank", match_col, match_cat_col,
               "R_c", "index_in_fasttext", "match_in_fasttext"]
    col_widths = [28, 22, 6, 28, 22, 12, 16, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    row_num = 2
    for index_word, matches in sorted(groups.items()):
        top = sorted(matches, key=lambda x: x["score"], reverse=True)[:top_n]
        in_ft = index_in_ft.get(index_word, False) if index_in_ft else "N/A"
        
        for rank, m in enumerate(top, 1):
            values = [
                index_word, m["index_cat"], rank, m["match_word"],
                m["match_cat"], round(m["score"], 6),
                in_ft, m.get("match_in_ft", "N/A")
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = get_fill(m["score"])
                cell.font = FONT_BODY
                cell.alignment = Alignment(
                    horizontal="center" if col_idx in (3, 7, 8) else "left"
                )
            row_num += 1

    wb.save(output_path)
    print(f"  Written: {output_path}  ({row_num - 2} rows)")

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Semantic Alignment Pipeline (Liu et al., 2025)"
    )
    parser.add_argument("--top", type=int, default=10,
                        help="Number of top matches per word (default 10)")
    args = parser.parse_args()

    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    for path, name in [
        (POLISH_CDI, "Polish CDI"),
        (ENGLISH_CDI, "English CDI"),
        (DICTIONARY, "Dictionary"),
        (FASTTEXT_PL, "Polish fastText"),
        (FASTTEXT_EN, "English fastText"),
    ]:
        if not os.path.exists(path):
            print(f"ERROR: {name} not found: {path}")
            sys.exit(1)

    run_pipeline(top_n=args.top)


if __name__ == "__main__":
    main()