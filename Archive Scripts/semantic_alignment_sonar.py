"""
semantic_alignment_sonar.py
============================
Combined approach:
  - Paper 2 (Hämmerl et al., 2022): SONAR for robust multilingual embeddings
  - Paper 1 (Liu et al., 2025): Similarity profiles + Pearson R_c for scoring

SONAR uses subword tokenization, so multi-word phrases, diminutives,
and sound effects all get vectors — no missing words.

Usage:
  python semantic_alignment_sonar.py --top 10
"""

import argparse
import csv
import os
import re
import sys
from collections import defaultdict

import numpy as np
from scipy.stats import pearsonr

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────

BASE_DIR = r"C:\Users\CAyre\Documents\Coding\TranslationClassification\TranslationClassification"
DATA_DIR = os.path.join(BASE_DIR, "data")
RESULTS_DIR = os.path.join(BASE_DIR, "results")

POLISH_CDI = os.path.join(DATA_DIR, "cdi_pl_diminutives.csv")
ENGLISH_CDI = os.path.join(DATA_DIR, "american_english_itemdata.csv")
DICTIONARY = os.path.join(DATA_DIR, "pl_en_dict.csv")

# Unique output names so nothing gets overwritten
TIMESTAMP = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_CSV = os.path.join(DATA_DIR, f"sonar_alignment_scores_{TIMESTAMP}.csv")
OUTPUT_EN_FOR_PL = os.path.join(RESULTS_DIR, f"top10_english_for_polish_sonar_{TIMESTAMP}.xlsx")
OUTPUT_PL_FOR_EN = os.path.join(RESULTS_DIR, f"top10_polish_for_english_sonar_{TIMESTAMP}.xlsx")

# ── Excel styles ──────────────────────────────────────────────────────────────

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BODY = Font(name="Arial", size=10)


def get_fill(score):
    if score >= 0.75:
        return FILL_GREEN
    elif score >= 0.60:
        return FILL_YELLOW
    else:
        return FILL_RED


# ── helpers ───────────────────────────────────────────────────────────────────

def strip_parens(text):
    return re.sub(r"\s*\(.*?\)", "", text).strip()


def load_cdi(path):
    seen, items = set(), []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            word = strip_parens(row["item_definition"])
            cat = row.get("category", "")
            if word not in seen:
                seen.add(word)
                items.append((word, cat))
    return items


def load_dictionary(path):
    mapping = defaultdict(list)
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            en_word = strip_parens(row["english"]).lower()
            pl_word = strip_parens(row["polish"]).lower()
            mapping[en_word].append(pl_word)
            if row.get("polish_alternatives"):
                for alt in row["polish_alternatives"].split("|"):
                    alt = strip_parens(alt).lower()
                    if alt and alt != pl_word:
                        mapping[en_word].append(alt)

    associates = []
    for en_word, pl_words in mapping.items():
        associates.append({"en": en_word, "pl_list": list(set(pl_words))})
    return associates


# ── SONAR embedding ──────────────────────────────────────────────────────────

def embed_words_sonar(words, lang_code):
    """
    Embed words using SONAR.
    Uses template sentences per Paper 2 insight (context improves representations).
    Returns dict: {word: normalized_vector}
    """
    from sonar.inference_pipelines.text import TextToEmbeddingModelPipeline

    print(f"  Loading SONAR for {lang_code}...")
    model = TextToEmbeddingModelPipeline(
        encoder="text_sonar_basic_encoder",
        tokenizer="text_sonar_basic_encoder"
    )

    # Template sentences in Polish and English
    if lang_code == "pl":
        templates = [
            "To jest {word}.",
            "Widzę {word}.",
            "Gdzie jest {word}?",
        ]
    else:
        templates = [
            "This is a {word}.",
            "I see a {word}.",
            "Where is the {word}?",
        ]

    word_vecs = {}
    batch_size = 32
    words_list = list(words)

    for i in range(0, len(words_list), batch_size):
        batch = words_list[i:i + batch_size]
        # For each word, create sentences with templates
        all_sentences = []
        word_indices = []
        for idx, word in enumerate(batch):
            for t in templates:
                all_sentences.append(t.format(word=word))
                word_indices.append(idx)

        # Embed all sentences at once
        embeddings = model.predict(all_sentences)

        # Average embeddings per word across templates
        for idx, word in enumerate(batch):
            idx_mask = [j for j, w_idx in enumerate(word_indices) if w_idx == idx]
            avg_vec = np.mean([embeddings[j] for j in idx_mask], axis=0)
            avg_vec = avg_vec / np.linalg.norm(avg_vec)
            word_vecs[word] = avg_vec

        if (i + batch_size) % 200 == 0:
            print(f"    Embedded {min(i + batch_size, len(words_list))}/{len(words_list)} words...")

    print(f"    Embedded {len(word_vecs)}/{len(words)} words")
    return word_vecs


# ── core algorithm ───────────────────────────────────────────────────────────

def build_associate_vectors(associates, pl_vecs, en_vecs):
    """
    For each shared associate, get its English vector and the average
    of its Polish equivalent vectors (Formula 1 from Liu et al., 2025).
    """
    en_vec_list = []
    pl_vec_list = []

    for assoc in associates:
        en_word = assoc["en"].lower()
        pl_words = [w.lower() for w in assoc["pl_list"]]

        en_vec = en_vecs.get(en_word)
        if en_vec is None:
            continue

        pl_vecs_for_word = []
        for pw in pl_words:
            if pw in pl_vecs:
                pl_vecs_for_word.append(pl_vecs[pw])

        if not pl_vecs_for_word:
            continue

        pl_vec = np.mean(pl_vecs_for_word, axis=0)
        pl_vec = pl_vec / (np.linalg.norm(pl_vec) + 1e-10)

        en_vec_list.append(en_vec)
        pl_vec_list.append(pl_vec)

    print(f"  Valid shared associates: {len(en_vec_list)}/{len(associates)}")
    return np.array(en_vec_list), np.array(pl_vec_list)


def semantic_alignment_score(pl_word, en_word, pl_vecs, en_vecs,
                              en_assoc_vecs, pl_assoc_vecs):
    """Calculate Pearson R_c between two words' similarity profiles."""
    pl_vec = pl_vecs.get(pl_word.lower())
    en_vec = en_vecs.get(en_word.lower())

    if pl_vec is None or en_vec is None:
        return None

    sims_pl = pl_assoc_vecs @ pl_vec
    sims_en = en_assoc_vecs @ en_vec

    if len(sims_pl) < 3:
        return None

    r_c, _ = pearsonr(sims_en, sims_pl)
    return r_c


# ── main pipeline ─────────────────────────────────────────────────────────────

def run_pipeline(top_n=10):
    # 1. Load data
    print("=" * 60)
    print("STEP 1: Loading CDI data and dictionary")
    print("=" * 60)

    pl_items = load_cdi(POLISH_CDI)
    en_items = load_cdi(ENGLISH_CDI)
    associates = load_dictionary(DICTIONARY)

    pl_words = [w for w, _ in pl_items]
    en_words = [w for w, _ in en_items]

    print(f"  Polish CDI words: {len(pl_words)}")
    print(f"  English CDI words: {len(en_words)}")
    print(f"  Dictionary entries: {len(associates)}")
    print(f"  Total pairs to score: {len(pl_words) * len(en_words):,}\n")

    # 2. Embed with SONAR
    print("=" * 60)
    print("STEP 2: Embedding words with SONAR")
    print("=" * 60)

    # Gather all words we need vectors for
    dict_pl_words = []
    for a in associates:
        dict_pl_words.extend(a["pl_list"])
    dict_en_words = [a["en"] for a in associates]

    all_pl = list(set(pl_words + dict_pl_words))
    all_en = list(set(en_words + dict_en_words))

    print(f"  Polish words to embed: {len(all_pl)}")
    print(f"  English words to embed: {len(all_en)}\n")

    pl_vecs = embed_words_sonar(all_pl, "pl")
    en_vecs = embed_words_sonar(all_en, "en")

    # Report any missing (should be none with SONAR)
    pl_missing = [w for w in pl_words if w.lower() not in pl_vecs]
    en_missing = [w for w in en_words if w.lower() not in en_vecs]

    if pl_missing:
        print(f"\n  ⚠ Polish CDI words missing: {len(pl_missing)}")
        print(f"    Examples: {pl_missing[:10]}")
    else:
        print(f"\n  ✓ All {len(pl_words)} Polish CDI words embedded")

    if en_missing:
        print(f"  ⚠ English CDI words missing: {len(en_missing)}")
        print(f"    Examples: {en_missing[:10]}")
    else:
        print(f"  ✓ All {len(en_words)} English CDI words embedded")

    # 3. Build associate vectors
    print("\n" + "=" * 60)
    print("STEP 3: Building shared associate vectors")
    print("=" * 60)

    en_assoc_vecs, pl_assoc_vecs = build_associate_vectors(
        associates, pl_vecs, en_vecs
    )

    if len(en_assoc_vecs) < 3:
        print("ERROR: Too few valid shared associates.")
        return

    # 4. Compute semantic alignment scores
    print("\n" + "=" * 60)
    print("STEP 4: Computing semantic alignment scores (Pearson R_c)")
    print("=" * 60)

    rows = []
    total = len(pl_words) * len(en_words)
    count = 0

    for pl_w, pl_cat in pl_items:
        for en_w, en_cat in en_items:
            r_c = semantic_alignment_score(
                pl_w, en_w, pl_vecs, en_vecs,
                en_assoc_vecs, pl_assoc_vecs
            )
            if r_c is not None and not np.isnan(r_c):
                rows.append({
                    "polish_word": pl_w,
                    "english_word": en_w,
                    "polish_category": pl_cat,
                    "english_category": en_cat,
                    "R_c": round(r_c, 6),
                })
            count += 1
            if count % 10000 == 0:
                print(f"  Processed {count}/{total} pairs...")

    print(f"  Scored: {len(rows)}, Skipped: {total - len(rows)}")

    # 5. Save CSV
    print(f"\nWriting {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "polish_word", "english_word", "polish_category",
            "english_category", "R_c"
        ])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    # 6. Excel output
    print("\n" + "=" * 60)
    print("STEP 5: Building Excel output")
    print("=" * 60)

    pl_groups = defaultdict(list)
    en_groups = defaultdict(list)

    for r in rows:
        pl_groups[r["polish_word"]].append({
            "match_word": r["english_word"],
            "match_cat": r["english_category"],
            "index_cat": r["polish_category"],
            "score": r["R_c"],
        })
        en_groups[r["english_word"]].append({
            "match_word": r["polish_word"],
            "match_cat": r["polish_category"],
            "index_cat": r["english_category"],
            "score": r["R_c"],
        })

    write_excel(pl_groups, top_n, OUTPUT_EN_FOR_PL,
                "polish_word", "polish_category",
                "english_word", "english_category")

    write_excel(en_groups, top_n, OUTPUT_PL_FOR_EN,
                "english_word", "english_category",
                "polish_word", "polish_category")

    # 7. Print top pairs
    rows_sorted = sorted(rows, key=lambda r: r["R_c"], reverse=True)
    print(f"\nTop 25 most semantically aligned pairs (SONAR + Paper 1):")
    print(f"  {'Polish':<30} {'English':<30} {'R_c':>10}")
    print("  " + "-" * 75)
    for r in rows_sorted[:25]:
        print(f"  {r['polish_word']:<30} {r['english_word']:<30} {r['R_c']:>10.4f}")

    print(f"\nAll done!")
    print(f"  {OUTPUT_CSV}")
    print(f"  {OUTPUT_EN_FOR_PL}")
    print(f"  {OUTPUT_PL_FOR_EN}")


def write_excel(groups, top_n, output_path, index_col, index_cat_col,
                match_col, match_cat_col):
    wb = Workbook()
    ws = wb.active
    ws.title = "Top matches"

    headers = [index_col, index_cat_col, "rank", match_col, match_cat_col, "R_c"]
    col_widths = [30, 22, 6, 30, 22, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    row_num = 2
    for index_word, matches in sorted(groups.items()):
        top = sorted(matches, key=lambda x: x["score"], reverse=True)[:top_n]
        for rank, m in enumerate(top, 1):
            values = [index_word, m["index_cat"], rank, m["match_word"],
                      m["match_cat"], round(m["score"], 6)]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = get_fill(m["score"])
                cell.font = FONT_BODY
                cell.alignment = Alignment(
                    horizontal="center" if col_idx == 3 else "left"
                )
            row_num += 1

    wb.save(output_path)
    print(f"  Written: {output_path}  ({row_num - 2} rows)")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Semantic Alignment with SONAR (Papers 1 + 2 combined)"
    )
    parser.add_argument("--top", type=int, default=10,
                        help="Number of top matches per word (default 10)")
    args = parser.parse_args()

    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    for path, name in [
        (POLISH_CDI, "Polish CDI"),
        (ENGLISH_CDI, "English CDI"),
        (DICTIONARY, "Dictionary"),
    ]:
        if not os.path.exists(path):
            print(f"ERROR: {name} not found: {path}")
            sys.exit(1)

    print(f"Output files will use timestamp: {TIMESTAMP}")
    print(f"  CSV:  {os.path.basename(OUTPUT_CSV)}")
    print(f"  Excel: {os.path.basename(OUTPUT_EN_FOR_PL)}")
    print(f"  Excel: {os.path.basename(OUTPUT_PL_FOR_EN)}")
    print()

    run_pipeline(top_n=args.top)


if __name__ == "__main__":
    main()