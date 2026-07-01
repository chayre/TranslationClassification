"""
semantic_alignment_categories.py
=================================
Uses CDI categories as the shared associate set instead of a word dictionary.

For each CDI category (Animals, Food, Body Parts, etc.), we compute:
  - The centroid of all Polish words in that category
  - The centroid of all English words in that category

These category centroids become the shared associates.
A word's similarity profile is its cosine similarity to each category centroid.
Pearson R_c measures how similarly two words relate to the category structure.

Usage:
  python semantic_alignment_categories.py --top 10
"""

import argparse
import csv
import os
import re
import sys
from collections import defaultdict

import numpy as np
from scipy.stats import pearsonr
from sentence_transformers import SentenceTransformer

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────

BASE_DIR = r"C:\Users\CAyre\Documents\Coding\TranslationClassification\TranslationClassification"
DATA_DIR = os.path.join(BASE_DIR, "data")
RESULTS_DIR = os.path.join(BASE_DIR, "results")

POLISH_CDI = os.path.join(DATA_DIR, "cdi_pl_diminutives.csv")
ENGLISH_CDI = os.path.join(DATA_DIR, "american_english_itemdata.csv")

OUTPUT_CSV = os.path.join(DATA_DIR, "category_alignment_scores.csv")
OUTPUT_EN_FOR_PL = os.path.join(RESULTS_DIR, "top10_english_for_polish_categories.xlsx")
OUTPUT_PL_FOR_EN = os.path.join(RESULTS_DIR, "top10_polish_for_english_categories.xlsx")

# ── Excel styles ──────────────────────────────────────────────────────────────

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BODY = Font(name="Arial", size=10)


def get_fill(score):
    if score >= 0.75:
        return FILL_GREEN
    elif score >= 0.60:
        return FILL_YELLOW
    else:
        return FILL_RED


# ── helpers ───────────────────────────────────────────────────────────────────

def strip_parens(text):
    return re.sub(r"\s*\(.*?\)", "", text).strip()


def load_cdi(path):
    """Load CDI words with their categories."""
    items = []
    seen = set()
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            word = strip_parens(row["item_definition"])
            cat = row.get("category", "").strip()
            if word and word not in seen:
                seen.add(word)
                items.append((word, cat))
    return items


# ── LaBSE embedding ──────────────────────────────────────────────────────────

def embed_words_labse(words):
    """Embed words using LaBSE with template sentences."""
    print("  Loading LaBSE model...")
    model = SentenceTransformer("sentence-transformers/LaBSE")
    print("  Model ready.\n")

    templates = [
        "This is a {word}.",
        "I see a {word}.",
        "Where is the {word}?",
    ]

    words_list = list(words)
    all_sentences = []
    word_to_indices = defaultdict(list)

    for word in words_list:
        for t in templates:
            word_to_indices[word].append(len(all_sentences))
            all_sentences.append(t.format(word=word))

    print(f"  Embedding {len(all_sentences)} sentences...")
    embeddings = model.encode(
        all_sentences,
        normalize_embeddings=True,
        show_progress_bar=True,
        convert_to_numpy=True
    )

    word_vecs = {}
    for word in words_list:
        indices = word_to_indices[word]
        avg_vec = np.mean(embeddings[indices], axis=0)
        avg_vec = avg_vec / (np.linalg.norm(avg_vec) + 1e-10)
        word_vecs[word] = avg_vec

    print(f"  Embedded {len(word_vecs)}/{len(words_list)} words")
    return word_vecs


# ── category centroids ───────────────────────────────────────────────────────

def build_category_centroids(items, vecs, language_name):
    """
    Group words by category and compute centroid vectors.
    Returns: dict of {category_name: centroid_vector}, list of category names
    """
    cat_words = defaultdict(list)
    for word, cat in items:
        vec = vecs.get(word.lower())
        if vec is not None and cat:
            cat_words[cat].append(vec)
    
    centroids = {}
    for cat, vec_list in cat_words.items():
        if len(vec_list) >= 2:  # Need at least 2 words for a meaningful centroid
            centroid = np.mean(vec_list, axis=0)
            centroid = centroid / (np.linalg.norm(centroid) + 1e-10)
            centroids[cat] = centroid
    
    print(f"  {language_name} categories: {len(centroids)} (from {len(cat_words)} total)")
    
    # Print category sizes
    for cat in sorted(centroids.keys()):
        print(f"    {cat}: {len(cat_words[cat])} words")
    
    return centroids


# ── core algorithm ──────────────────────────────────────────────────────────

def semantic_alignment_score(pl_word, en_word, pl_vecs, en_vecs,
                              en_centroids, pl_centroids, shared_categories):
    """
    Calculate Pearson R_c using category centroids as shared associates.
    Each word's similarity profile is its cosine similarity to each shared category.
    """
    pl_vec = pl_vecs.get(pl_word.lower())
    en_vec = en_vecs.get(en_word.lower())

    if pl_vec is None or en_vec is None:
        return None

    sims_pl = []
    sims_en = []

    for cat in shared_categories:
        pl_centroid = pl_centroids.get(cat)
        en_centroid = en_centroids.get(cat)
        
        if pl_centroid is not None and en_centroid is not None:
            sims_pl.append(np.dot(pl_centroid, pl_vec))
            sims_en.append(np.dot(en_centroid, en_vec))

    if len(sims_pl) < 3:
        return None

    r_c, _ = pearsonr(sims_en, sims_pl)
    return r_c


# ── main pipeline ─────────────────────────────────────────────────────────────

def run_pipeline(top_n=10):
    # 1. Load data
    print("=" * 60)
    print("STEP 1: Loading CDI data")
    print("=" * 60)

    pl_items = load_cdi(POLISH_CDI)
    en_items = load_cdi(ENGLISH_CDI)

    pl_words = [w for w, _ in pl_items]
    en_words = [w for w, _ in en_items]

    # Count categories
    pl_cats = set(cat for _, cat in pl_items if cat)
    en_cats = set(cat for _, cat in en_items if cat)
    shared_cats = sorted(pl_cats & en_cats)

    print(f"  Polish CDI words: {len(pl_words)}")
    print(f"  English CDI words: {len(en_words)}")
    print(f"  Polish categories: {len(pl_cats)}")
    print(f"  English categories: {len(en_cats)}")
    print(f"  Shared categories: {len(shared_cats)}")
    print(f"  Shared categories: {shared_cats}")
    print(f"  Total pairs to score: {len(pl_words) * len(en_words):,}")

    # 2. Embed with LaBSE
    print("\n" + "=" * 60)
    print("STEP 2: Embedding all CDI words with LaBSE")
    print("=" * 60)

    all_words = list(set(pl_words + en_words))
    print(f"  Total unique words: {len(all_words)}\n")

    all_vecs = embed_words_labse(all_words)

    pl_vecs = {w.lower(): all_vecs[w] for w in pl_words if w in all_vecs}
    en_vecs = {w.lower(): all_vecs[w] for w in en_words if w in all_vecs}

    print(f"\n  ✓ All {len(pl_words)} Polish words embedded")
    print(f"  ✓ All {len(en_words)} English words embedded")

    # 3. Build category centroids
    print("\n" + "=" * 60)
    print("STEP 3: Building category centroids")
    print("=" * 60)

    pl_centroids = build_category_centroids(pl_items, pl_vecs, "Polish")
    en_centroids = build_category_centroids(en_items, en_vecs, "English")

    # Keep only categories present in both languages
    valid_cats = sorted(set(pl_centroids.keys()) & set(en_centroids.keys()))
    print(f"\n  Valid shared categories: {len(valid_cats)}")
    print(f"  Categories: {valid_cats}")

    if len(valid_cats) < 3:
        print("ERROR: Need at least 3 shared categories for Pearson correlation.")
        return

    # 4. Compute semantic alignment scores
    print("\n" + "=" * 60)
    print("STEP 4: Computing semantic alignment scores (Pearson R_c)")
    print("=" * 60)
    print(f"  Using {len(valid_cats)} category centroids as shared associates")

    rows = []
    total = len(pl_words) * len(en_words)
    count = 0

    for pl_w, pl_cat in pl_items:
        for en_w, en_cat in en_items:
            r_c = semantic_alignment_score(
                pl_w, en_w, pl_vecs, en_vecs,
                en_centroids, pl_centroids, valid_cats
            )
            if r_c is not None and not np.isnan(r_c):
                rows.append({
                    "polish_word": pl_w,
                    "english_word": en_w,
                    "polish_category": pl_cat,
                    "english_category": en_cat,
                    "R_c": round(r_c, 6),
                })
            count += 1
            if count % 10000 == 0:
                print(f"  Processed {count}/{total} pairs...")

    print(f"  Scored: {len(rows)}, Skipped: {total - len(rows)}")

    # 5. Save CSV
    print(f"\nWriting {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "polish_word", "english_word", "polish_category",
            "english_category", "R_c"
        ])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    # 6. Excel output
    print("\n" + "=" * 60)
    print("STEP 5: Building Excel output")
    print("=" * 60)

    pl_groups = defaultdict(list)
    en_groups = defaultdict(list)

    for r in rows:
        pl_groups[r["polish_word"]].append({
            "match_word": r["english_word"],
            "match_cat": r["english_category"],
            "index_cat": r["polish_category"],
            "score": r["R_c"],
        })
        en_groups[r["english_word"]].append({
            "match_word": r["polish_word"],
            "match_cat": r["polish_category"],
            "index_cat": r["english_category"],
            "score": r["R_c"],
        })

    write_excel(pl_groups, top_n, OUTPUT_EN_FOR_PL,
                "polish_word", "polish_category",
                "english_word", "english_category")

    write_excel(en_groups, top_n, OUTPUT_PL_FOR_EN,
                "english_word", "english_category",
                "polish_word", "polish_category")

    # 7. Print top pairs
    rows_sorted = sorted(rows, key=lambda r: r["R_c"], reverse=True)
    print(f"\nTop 25 most semantically aligned pairs (Category Centroids):")
    print(f"  {'Polish':<30} {'English':<30} {'R_c':>10}")
    print("  " + "-" * 75)
    for r in rows_sorted[:25]:
        print(f"  {r['polish_word']:<30} {r['english_word']:<30} {r['R_c']:>10.4f}")

    # Check frog specifically
    frog_rows = [r for r in rows if r["polish_word"].lower() == "żaba" or r["english_word"].lower() == "frog"]
    if frog_rows:
        print(f"\n  ── 'frog' / 'żaba' matches ──")
        frog_sorted = sorted(frog_rows, key=lambda r: r["R_c"], reverse=True)
        for r in frog_sorted[:5]:
            print(f"    {r['polish_word']:<30} {r['english_word']:<30} {r['R_c']:>10.4f}")

    print(f"\nAll done!")
    print(f"  Scores: {OUTPUT_CSV}")
    print(f"  Excel:  {OUTPUT_EN_FOR_PL}")
    print(f"  Excel:  {OUTPUT_PL_FOR_EN}")


def write_excel(groups, top_n, output_path, index_col, index_cat_col,
                match_col, match_cat_col):
    wb = Workbook()
    ws = wb.active
    ws.title = "Top matches"

    headers = [index_col, index_cat_col, "rank", match_col, match_cat_col, "R_c"]
    col_widths = [30, 22, 6, 30, 22, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    row_num = 2
    for index_word, matches in sorted(groups.items()):
        top = sorted(matches, key=lambda x: x["score"], reverse=True)[:top_n]
        for rank, m in enumerate(top, 1):
            values = [index_word, m["index_cat"], rank, m["match_word"],
                      m["match_cat"], round(m["score"], 6)]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = get_fill(m["score"])
                cell.font = FONT_BODY
                cell.alignment = Alignment(
                    horizontal="center" if col_idx == 3 else "left"
                )
            row_num += 1

    wb.save(output_path)
    print(f"  Written: {output_path}  ({row_num - 2} rows)")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Semantic Alignment with Category Centroids"
    )
    parser.add_argument("--top", type=int, default=10,
                        help="Number of top matches per word (default 10)")
    args = parser.parse_args()

    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    for path, name in [
        (POLISH_CDI, "Polish CDI"),
        (ENGLISH_CDI, "English CDI"),
    ]:
        if not os.path.exists(path):
            print(f"ERROR: {name} not found: {path}")
            sys.exit(1)

    print("Semantic Alignment Pipeline (Category Centroids)")
    print("=" * 50)
    print()
    print("Output files:")
    print(f"  Scores: {OUTPUT_CSV}")
    print(f"  Excel:  {OUTPUT_EN_FOR_PL}")
    print(f"  Excel:  {OUTPUT_PL_FOR_EN}")
    print()

    run_pipeline(top_n=args.top)


if __name__ == "__main__":
    main()