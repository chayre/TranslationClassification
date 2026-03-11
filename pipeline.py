"""
CDI Cross-Linguistic Semantic Similarity Pipeline
==================================================
Full pipeline from raw CDI data to colour-coded Excel output.

Folder structure:
  data/
    norwegian_itemdata.csv
    american_english_itemdata.csv
    unilemma_pairs.csv
    semantic_distances.csv        (generated, can be reused with --skip-embedding)
  results/
    top10_english_for_norwegian.xlsx
    top10_norwegian_for_english.xlsx

Steps:
  1. Embed all Norwegian and English CDI words using LaBSE
  2. Compute all-vs-all cosine similarity matrix
  3. Flag human-verified translation equivalents (unilemmas from Wordbank)
  4. Check string-similar non-unilemma pairs for false cognates via LLM
  5. Build top-10 match tables and write colour-coded Excel files

Scoring:
  - Unilemma-verified pair          -> score = 1.0
  - LLM-confirmed false cognate     -> score = 0.0
  - Everything else                 -> score = LaBSE cosine similarity

Colour coding:
  Green  (score >= 0.75) : strong match
  Yellow (score 0.60-0.75): semantically related
  Red    (score < 0.60)  : weak / unrelated / false cognate

Setup:
  pip install -r requirements.txt

Run (full pipeline):
  python pipeline.py

Run (skip LaBSE re-embedding, reuse existing semantic_distances.csv):
  python pipeline.py --skip-embedding

Run (skip LLM false-cognate check):
  python pipeline.py --no-llm

Run (both skips, just rebuild Excel from existing data):
  python pipeline.py --skip-embedding --no-llm
"""

import argparse
import csv
import json
import os
import re
import time
from collections import defaultdict

import numpy as np

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── folder paths ──────────────────────────────────────────────────────────────

DATA_DIR    = "data"
RESULTS_DIR = "results"

POLISH_CSV   = os.path.join(DATA_DIR, "cdi_pl_diminutives.csv")
ENGLISH_CSV     = os.path.join(DATA_DIR, "american_english_itemdata.csv")
UNILEMMA_CSV    = os.path.join(DATA_DIR, "unilemma_pairs.csv")
DISTANCES_CSV   = os.path.join(DATA_DIR, "semantic_distances.csv")

OUTPUT_EN_FOR_PL = os.path.join(RESULTS_DIR, "top10_english_for_polish.xlsx")
OUTPUT_PL_FOR_EN = os.path.join(RESULTS_DIR, "top10_polish_for_english.xlsx")

# ── Excel styles ──────────────────────────────────────────────────────────────

FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_ORANGE = PatternFill("solid", fgColor="FFCCCC")
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BODY   = Font(name="Arial", size=10)


# ── helpers ───────────────────────────────────────────────────────────────────

def strip_parens(text):
    return re.sub(r"\s*\(.*?\)", "", text).strip()


def levenshtein_sim(a, b):
    a, b = a.lower(), b.lower()
    m, n = len(a), len(b)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[:], i
        for j in range(1, n + 1):
            dp[j] = prev[j-1] if a[i-1] == b[j-1] else 1 + min(prev[j-1], prev[j], dp[j-1])
    return 1 - dp[n] / max(m, n, 1)


def get_fill(score):
    if score >= 0.75:
        return FILL_GREEN
    elif score >= 0.60:
        return FILL_YELLOW
    else:
        return FILL_RED


# ── data loading ──────────────────────────────────────────────────────────────

def load_cdi(path):
    seen, items = set(), []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            word = strip_parens(row["item_definition"])
            cat  = row["category"]
            if word not in seen:
                seen.add(word)
                items.append((word, cat))
    return items


def load_unilemmas(path):
    mapping = {}
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        # Check the actual columns in the CSV
        columns = reader.fieldnames
        lang_a_col = "polish_word" if "polish_word" in columns else "norwegian_word" if "norwegian_word" in columns else None
        if lang_a_col is None:
            print(f"Warning: No suitable language A column found in {path} (expected 'polish_word' or 'norwegian_word'). Skipping unilemmas.")
            return {}
        print(f"Using column '{lang_a_col}' for language A words from {path}.")
        for row in reader:
            pl_w = strip_parens(row[lang_a_col])
            en_w = strip_parens(row["english_word"])
            mapping[(pl_w, en_w)] = row["uni_lemma"].strip()
    return mapping


def load_distances(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            is_unilemma = row.get("is_unilemma_match", "False").strip().lower() == "true"
            cosine = float(row["cosine_sim"])
            rows.append({
                "polish_word":     row["polish_word"],
                "english_word":       row["english_word"],
                "polish_category": row.get("polish_category", ""),
                "english_category":   row.get("english_category", ""),
                "cosine_sim":         cosine,
                "is_unilemma_match":  is_unilemma,
                "uni_lemma":          row.get("uni_lemma", ""),
                "score":              1.0 if is_unilemma else cosine,
                "false_cognate":      False,
            })
    return rows


# ── step 1+2+3: embed and compute distances ───────────────────────────────────

def load_diminutives(path):
    """Returns a set of Polish words marked as diminutive."""
    diminutives = set()
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("diminutive", "").strip().lower() == "yes":
                diminutives.add(strip_parens(row["item_definition"]))
    return diminutives


def embed_labse(words_a, words_b):
    from sentence_transformers import SentenceTransformer
    print("Loading LaBSE model (downloads ~1.9 GB on first run)...")
    model = SentenceTransformer("sentence-transformers/LaBSE")
    print("Model ready.\n")
    print("Embedding language A words...")
    vecs_a = model.encode(words_a, normalize_embeddings=True,
                          show_progress_bar=True, convert_to_numpy=True)
    print("Embedding language B words...")
    vecs_b = model.encode(words_b, normalize_embeddings=True,
                          show_progress_bar=True, convert_to_numpy=True)
    return vecs_a, vecs_b


def embed_fasttext(words_a, words_b, vec_path_a, vec_path_b):
    from gensim.models import KeyedVectors
    import numpy as np

    def load_vecs(path, words):
        print(f"  Loading aligned vectors from {path}...")

        # First try gensim's standard loader (fast & handles header/binary)
        try:
            model = KeyedVectors.load_word2vec_format(
                path,
                binary=False,              # text format (.vec)
                unicode_errors='ignore',
                limit=None
            )
            print(f"    Loaded {len(model)} words, dimension {model.vector_size} (via gensim loader)")
            word_vec = {w: model[w] for w in model}  # dict for easy lookup
            dim = model.vector_size
            invalid_lines = 0  # no manual parsing needed if this succeeds

        except Exception as e:
            print(f"    Gensim loader failed ({e}), falling back to manual parsing...")
            # Manual fallback for corrupted/invalid lines
            word_vec = {}
            invalid_lines = 0
            with open(path, 'r', encoding='utf-8', errors='ignore') as fin:
                first_line = fin.readline().strip()
                try:
                    vocab_size, vec_dim = map(int, first_line.split())
                    print(f"    Header: {vocab_size} words, dim {vec_dim}")
                except:
                    print("    No valid header → assuming no header")
                    fin.seek(0)
                    vec_dim = None

                for line in fin:
                    if not line.strip():
                        continue
                    parts = line.rstrip().split(' ')
                    if len(parts) < 2:
                        invalid_lines += 1
                        continue
                    word = parts[0]
                    try:
                        vec = [float(x) for x in parts[1:]]
                        expected_dim = vec_dim or 300
                        if len(vec) != expected_dim:
                            invalid_lines += 1
                            continue
                        word_vec[word] = np.array(vec, dtype=np.float32)
                    except ValueError:
                        invalid_lines += 1
                        continue

            if not word_vec:
                raise ValueError(f"No valid vectors loaded from {path}")
            dim = len(next(iter(word_vec.values())))
            print(f"    Loaded {len(word_vec)} valid words, dimension {dim} (manual parse)")

        if invalid_lines > 0:
            print(f"    Skipped {invalid_lines} invalid or corrupted lines")

        # Build normalized vectors for our CDI words
        vecs = []
        missing = 0
        for w in words:
            if w in word_vec:
                vecs.append(word_vec[w])
            else:
                missing += 1
                vecs.append(np.zeros(dim))
        if missing > 0:
            print(f"    Warning: {missing}/{len(words)} words not found → using zero vectors")

        vecs = np.array(vecs)
        norms = np.linalg.norm(vecs, axis=1, keepdims=True)
        norms[norms == 0] = 1
        return vecs / norms

    print("Embedding language A words...")
    vecs_a = load_vecs(vec_path_a, words_a)
    print("Embedding language B words...")
    vecs_b = load_vecs(vec_path_b, words_b)
    return vecs_a, vecs_b

def run_embedding(model_name="labse", fasttext_vec_a=None, fasttext_vec_b=None, top_n_print=25):

    unilemma_map = {}
    if os.path.exists(UNILEMMA_CSV):
        print(f"Loading unilemmas from {UNILEMMA_CSV}...")
        unilemma_map = load_unilemmas(UNILEMMA_CSV)
        print(f"  {len(unilemma_map)} verified pairs loaded.\n")
    else:
        print(f"Warning: {UNILEMMA_CSV} not found, skipping unilemma flags.\n")

    pl_items = load_cdi(POLISH_CSV)
    en_items = load_cdi(ENGLISH_CSV)
    pl_words = [w for w, _ in pl_items]
    en_words = [w for w, _ in en_items]

    print(f"Language A words: {len(pl_words)}")
    print(f"Language B words: {len(en_words)}")
    print(f"Total pairs:      {len(pl_words) * len(en_words):,}\n")

    if model_name == "labse":
        pl_vecs, en_vecs = embed_labse(pl_words, en_words)
    elif model_name == "fasttext":
        if not fasttext_vec_a or not fasttext_vec_b:
            print("ERROR: --model fasttext requires --fasttext-vec-a and --fasttext-vec-b")
            print("  Run download_fasttext.py first to get the vector files.")
            return
        pl_vecs, en_vecs = embed_fasttext(pl_words, en_words, fasttext_vec_a, fasttext_vec_b)
    else:
        print(f"ERROR: Unknown model '{model_name}'. Choose from: labse, fasttext")
        return

    print("\nComputing similarity matrix...")
    sim_matrix = pl_vecs @ en_vecs.T
    print("Done.\n")

    print(f"Writing {DISTANCES_CSV}...")
    unilemma_matches = 0
    fieldnames = ["polish_word", "english_word", "polish_category",
                  "english_category", "cosine_sim", "is_unilemma_match", "uni_lemma"]

    with open(DISTANCES_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for i, (pl_w, pl_cat) in enumerate(pl_items):
            for j, (en_w, en_cat) in enumerate(en_items):
                uni = unilemma_map.get((pl_w, en_w), "")
                is_match = bool(uni)
                if is_match:
                    unilemma_matches += 1
                writer.writerow({
                    "polish_word":     pl_w,
                    "english_word":       en_w,
                    "polish_category": pl_cat,
                    "english_category":   en_cat,
                    "cosine_sim":         round(float(sim_matrix[i, j]), 6),
                    "is_unilemma_match":  is_match,
                    "uni_lemma":          uni,
                })

    print(f"  Unilemma-flagged pairs: {unilemma_matches}")

    all_rows = [
        (pl_items[i][0], en_items[j][0], float(sim_matrix[i, j]),
         bool(unilemma_map.get((pl_items[i][0], en_items[j][0]), "")))
        for i in range(len(pl_items)) for j in range(len(en_items))
    ]
    all_rows.sort(key=lambda r: r[2], reverse=True)
    print(f"\nTop {top_n_print} most similar pairs:")
    print(f"  {'Language A':<25} {'Language B':<25} {'cosine_sim':>10}  unilemma")
    print("  " + "-" * 70)
    for pl_w, en_w, s, u in all_rows[:top_n_print]:
        print(f"  {pl_w:<25} {en_w:<25} {s:>10.4f}  {'*' if u else ''}")


# ── step 4: false cognate check ───────────────────────────────────────────────
'''
    system = "You are a Norwegian-English linguistics expert. Respond only with valid JSON, no extra text."
    user = (
        "Check these Norwegian-English word pairs for false cognates.\n"
        "A false cognate looks similar in both languages but has a different meaning.\n\n"
        "Known Norwegian-English false cognates:\n"
        "- bad (NO) = bath/bathtub, NOT the English word bad\n"
        "- her (NO) = here, NOT the English pronoun her\n"
        "- i (NO) = in (preposition), NOT the English pronoun I\n"
        "- man (NO) = one/someone (impersonal), NOT adult male\n"
        "- full (NO) = drunk, NOT not-empty\n"
        "- hit (NO) = found (past tense of hitte), NOT to strike\n"
        "- men (NO) = but (conjunction), NOT plural of man\n"
        "- sin (NO) = his/her/its (reflexive possessive), NOT wrongdoing\n\n"
        "Known TRUE matches: arm, ball, tiger, egg, glass, hammer, radio\n\n"
        "Return a JSON array with one object per pair:\n"
        '[{"norwegian":"...","english":"...","is_false_cognate":true}]\n\n'
        "Pairs to check:\n"
        + pairs_text
    )
'''
def check_false_cognates(client, pairs):
    if not pairs:
        return {}

    pairs_text = "\n".join(
        f"{i+1}. Polish='{pl}' | English='{en}'"
        for i, (pl, en) in enumerate(pairs)
    )

    system = "You are a Polish-English linguistics expert. Respond only with valid JSON, no extra text."

    user = (
        "Check these Polish-English word pairs for false cognates.\n"
        "A false cognate looks similar in both languages but has a different meaning.\n\n"
        "Known Polish-English false cognates:\n"
        "- aktualny (PL) = current/up-to-date, NOT actual\n"
        "- aktualnie (PL) = currently/presently, NOT actually\n"
        "- ale (PL) = but, NOT ale (drink)\n"
        "- but (PL) = shoe, NOT but (conjunction)\n"
        "- data (PL) = date (calendar), NOT data (information)\n"
        "- dywan (PL) = carpet, NOT divan (sofa)\n"
        "- fabryka (PL) = factory, NOT fabric\n"
        "- fartuch (PL) = apron, NOT fart (joke)\n"
        "- herbata (PL) = tea, NOT herb (plant)\n"
        "\n"
        "Known TRUE (cognate / similar meaning) matches: armia, bank, hotel, internet, kamera (in some contexts), radio, telewizja, tiger, university\n\n"
        "Return a JSON array with one object per pair:\n"
        '[{"polish":"...","english":"...","is_false_cognate":true}]\n\n'
        "Pairs to check:\n"
        + pairs_text
    )

    import anthropic
    for attempt in range(5):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=2048,
                system=system,
                messages=[{"role": "user", "content": user}]
            )
            break
        except Exception as e:
            if "overloaded" in str(e).lower() and attempt < 4:
                wait = 10 * (attempt + 1)
                print(f"    API overloaded, retrying in {wait}s... (attempt {attempt+1}/5)")
                time.sleep(wait)
            else:
                raise

    raw = response.content[0].text.strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw)
    raw = re.sub(r"\n?```$", "", raw)

    try:
        results = json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\[.*\]", raw, re.DOTALL)
        if match:
            try:
                results = json.loads(match.group(0))
            except json.JSONDecodeError:
                print(f"    WARNING: could not parse response, skipping batch")
                print(f"    Raw (first 500 chars): {raw[:500]}")
                return {}
        else:
            print(f"    WARNING: could not parse response, skipping batch")
            print(f"    Raw (first 500 chars): {raw[:500]}")
            return {}

    return {(r["polish"], r["english"]): r["is_false_cognate"] for r in results}


def apply_false_cognate_penalties(rows, api_key):
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)

    suspicious = list(set(
        (r["polish_word"], r["english_word"])
        for r in rows
        if not r["is_unilemma_match"]
        and levenshtein_sim(r["polish_word"], r["english_word"]) > 0.5
    ))
    print(f"  Suspicious string-similar pairs: {len(suspicious)}")

    batch_size = 50
    batches = [suspicious[i:i+batch_size] for i in range(0, len(suspicious), batch_size)]
    print(f"  Calling claude-sonnet in {len(batches)} batches...")

    false_cognate_map = {}
    for idx, batch in enumerate(batches, 1):
        print(f"    Batch {idx}/{len(batches)} ({len(batch)} pairs)...")
        false_cognate_map.update(check_false_cognates(client, batch))

    confirmed = sum(1 for v in false_cognate_map.values() if v)
    print(f"  False cognates confirmed: {confirmed}/{len(suspicious)}")

    for r in rows:
        key = (r["polish_word"], r["english_word"])
        if false_cognate_map.get(key):
            r["score"] = 0.0
            r["false_cognate"] = True

    return rows


# ── step 5: Excel output ──────────────────────────────────────────────────────

def write_excel(groups, top_n, output_path, index_col, index_cat_col, match_col, match_cat_col, diminutive_set=None):
    wb = Workbook()

    # Sheet 1: Top matches
    ws = wb.active
    ws.title = "Top matches"
    headers    = [index_col, index_cat_col, "rank", match_col, match_cat_col,
                  "score", "cosine_sim", "is_unilemma_match", "false_cognate", "uni_lemma"]
    col_widths = [28, 22, 6, 28, 22, 8, 11, 20, 14, 20]
    if diminutive_set is not None:
        headers.append("is_diminutive")
        col_widths.append(14)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    row_num = 2
    for index_word, matches in sorted(groups.items()):
        top = sorted(matches, key=lambda x: x["score"], reverse=True)[:top_n]
        for rank, m in enumerate(top, 1):
            values = [index_word, m["index_cat"], rank, m["match_word"], m["match_cat"],
                      round(m["score"], 6), round(m["cosine_sim"], 6),
                      m["is_unilemma_match"], m["false_cognate"], m["uni_lemma"]]
            if diminutive_set is not None:
                values.append(index_word in diminutive_set)
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = get_fill(m["score"])
                cell.font = FONT_BODY
                cell.alignment = Alignment(horizontal="center" if col_idx == 3 else "left")
            row_num += 1

    # Sheet 2: False cognates
    ws2 = wb.create_sheet("False Cognates")
    fc_headers    = [index_col, index_cat_col, match_col, match_cat_col, "cosine_sim"]
    fc_col_widths = [28, 22, 28, 22, 12]

    for col_idx, (header, width) in enumerate(zip(fc_headers, fc_col_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.freeze_panes = "A2"

    fc_row = 2
    for index_word, matches in sorted(groups.items()):
        for m in sorted([m for m in matches if m["false_cognate"]],
                        key=lambda x: x["cosine_sim"], reverse=True):
            values = [index_word, m["index_cat"], m["match_word"], m["match_cat"],
                      round(m["cosine_sim"], 6)]
            for col_idx, value in enumerate(values, start=1):
                cell = ws2.cell(row=fc_row, column=col_idx, value=value)
                cell.fill = FILL_ORANGE
                cell.font = FONT_BODY
                cell.alignment = Alignment(horizontal="left")
            fc_row += 1

    wb.save(output_path)
    print(f"  Written: {output_path}  ({row_num-2} match rows, {fc_row-2} false cognate rows)")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CDI semantic similarity pipeline")
    parser.add_argument("--skip-embedding",  action="store_true",
                        help="Skip embedding step, load existing semantic_distances.csv")
    parser.add_argument("--model",           default="labse",
                        choices=["labse", "fasttext"],
                        help="Embedding model to use (default: labse)")
    parser.add_argument("--fasttext-vec-a",  default=None,
                        help="Path to fastText aligned .vec file for language A")
    parser.add_argument("--fasttext-vec-b",  default=None,
                        help="Path to fastText aligned .vec file for language B")
    parser.add_argument("--no-llm",          action="store_true",
                        help="Skip LLM false-cognate check")
    parser.add_argument("--top",             type=int, default=10,
                        help="Number of top matches per word (default 10)")
    parser.add_argument("--api-key",         default=None,
                        help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    parser.add_argument("--mark-diminutives", action="store_true",
                        help="Add 'is_diminutive' column to Excel output for Polish words (from cdi_pl_diminutives.csv)")
    args = parser.parse_args()

    # Create output folders if they don't exist
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    # ── step 1-3: embedding ───────────────────────────────────────────────────
    if args.skip_embedding:
        if not os.path.exists(DISTANCES_CSV):
            print(f"Error: --skip-embedding set but {DISTANCES_CSV} not found.")
            return
        print(f"Skipping embedding, loading {DISTANCES_CSV}...")
    else:
        print("=" * 60)
        print(f"STEP 1-3: {args.model.upper()} embedding + unilemma flagging")
        print("=" * 60)
        run_embedding(
            model_name=args.model,
            fasttext_vec_a=args.fasttext_vec_a,
            fasttext_vec_b=args.fasttext_vec_b,
        )
        print()

    # ── step 4: false cognate check ───────────────────────────────────────────
    print("=" * 60)
    print("STEP 4: Loading distances" + (" + false cognate check" if not args.no_llm else ""))
    print("=" * 60)
    rows = load_distances(DISTANCES_CSV)
    print(f"  {len(rows):,} pairs loaded.")
    print(f"  Unilemma matches: {sum(1 for r in rows if r['is_unilemma_match'])}")

    if not args.no_llm:
        api_key = args.api_key or os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            print("\nWarning: no ANTHROPIC_API_KEY found, skipping false-cognate check.")
            print("Pass --api-key or set the ANTHROPIC_API_KEY environment variable.\n")
        else:
            print("\nChecking string-similar pairs for false cognates...")
            rows = apply_false_cognate_penalties(rows, api_key)
    print()

    # ── step 5: Excel output ──────────────────────────────────────────────────
    print("=" * 60)
    print("STEP 5: Building Excel output")
    print("=" * 60)
    diminutive_set = None
    if args.mark_diminutives:
        print(f"  Loading diminutives from {POLISH_CSV}...")
        diminutive_set = load_diminutives(POLISH_CSV)
        print(f"  {len(diminutive_set)} diminutive words loaded.")

    no_groups = defaultdict(list)
    en_groups = defaultdict(list)
    for r in rows:
        entry_no = {"match_word": r["english_word"],  "match_cat": r["english_category"],
                    "index_cat": r["polish_category"], "score": r["score"],
                    "cosine_sim": r["cosine_sim"], "is_unilemma_match": r["is_unilemma_match"],
                    "false_cognate": r["false_cognate"], "uni_lemma": r["uni_lemma"]}
        entry_en = {"match_word": r["polish_word"], "match_cat": r["polish_category"],
                    "index_cat": r["english_category"], "score": r["score"],
                    "cosine_sim": r["cosine_sim"], "is_unilemma_match": r["is_unilemma_match"],
                    "false_cognate": r["false_cognate"], "uni_lemma": r["uni_lemma"]}
        no_groups[r["polish_word"]].append(entry_no)
        en_groups[r["english_word"]].append(entry_en)

    print(f"Writing top-{args.top} English matches for each Polish word...")
    write_excel(groups=no_groups, top_n=args.top, output_path=OUTPUT_EN_FOR_PL,
                index_col="polish_word", index_cat_col="polish_category",
                match_col="english_word",   match_cat_col="english_category",
                diminutive_set=diminutive_set)

    print(f"Writing top-{args.top} Polish matches for each English word...")
    write_excel(groups=en_groups, top_n=args.top, output_path=OUTPUT_PL_FOR_EN,
                index_col="english_word",   index_cat_col="english_category",
                match_col="polish_word", match_cat_col="polish_category")

    print("\nAll done!")
    print(f"\nOutputs:")
    print(f"  {DISTANCES_CSV}")
    print(f"  {OUTPUT_EN_FOR_PL}")
    print(f"  {OUTPUT_PL_FOR_EN}")
    print("\nColour key:")
    print("  Green  (score >= 0.75) : strong match")
    print("  Yellow (score 0.60-0.75): semantically related")
    print("  Red    (score < 0.60)  : weak / unrelated / false cognate")


if __name__ == "__main__":
    main()