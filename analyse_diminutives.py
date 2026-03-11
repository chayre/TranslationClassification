"""
Diminutive vs Non-Diminutive Cosine Similarity Analysis
========================================================
Tests whether diminutive Polish words have significantly different
top-match cosine similarity scores compared to non-diminutive words.

Uses Mann-Whitney U test (non-parametric, appropriate for small groups).

Usage:
    python analyse_diminutives.py
    python analyse_diminutives.py --input results/top10_english_for_polish.xlsx
"""

import argparse
import os

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import numpy as np
import openpyxl
from scipy import stats


def load_data(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    diminutive, non_diminutive = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("rank") == 1:
            score = d.get("score", d.get("cosine_sim", 0)) or 0
            word = d.get("polish_word", "")
            is_dim = bool(d.get("is_diminutive", False))
            if is_dim:
                diminutive.append((word, score))
            else:
                non_diminutive.append((word, score))

    return diminutive, non_diminutive


def run_stats(dim_scores, nondim_scores):
    u_stat, p_value = stats.mannwhitneyu(dim_scores, nondim_scores, alternative="two-sided")
    n1, n2 = len(dim_scores), len(nondim_scores)
    # rank-biserial correlation as effect size
    r = 1 - (2 * u_stat) / (n1 * n2)
    return {
        "u_stat": u_stat,
        "p_value": p_value,
        "effect_size_r": r,
        "dim_mean": np.mean(dim_scores),
        "dim_median": np.median(dim_scores),
        "dim_std": np.std(dim_scores),
        "nondim_mean": np.mean(nondim_scores),
        "nondim_median": np.median(nondim_scores),
        "nondim_std": np.std(nondim_scores),
        "n_dim": n1,
        "n_nondim": n2,
    }


def plot(dim, nondim, results, output_path):
    dim_scores    = [s for _, s in dim]
    nondim_scores = [s for _, s in nondim]

    fig = plt.figure(figsize=(14, 10))
    fig.patch.set_facecolor("#F8F9FA")
    gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.4, wspace=0.35)

    BLUE   = "#4C78A8"
    ORANGE = "#F58518"

    # ── 1. Violin + strip plot ────────────────────────────────────────────────
    ax1 = fig.add_subplot(gs[0, 0])
    ax1.set_facecolor("#F8F9FA")

    parts = ax1.violinplot([nondim_scores, dim_scores], positions=[1, 2],
                           showmedians=True, showextrema=True)
    for i, (pc, col) in enumerate(zip(parts["bodies"], [BLUE, ORANGE])):
        pc.set_facecolor(col)
        pc.set_alpha(0.6)
    parts["cmedians"].set_color("black")
    parts["cmedians"].set_linewidth(2)
    for key in ["cbars", "cmins", "cmaxes"]:
        parts[key].set_color("grey")

    # Jittered points
    jitter = 0.07
    ax1.scatter(np.random.normal(1, jitter, len(nondim_scores)), nondim_scores,
                color=BLUE, alpha=0.3, s=12, zorder=3)
    ax1.scatter(np.random.normal(2, jitter, len(dim_scores)), dim_scores,
                color=ORANGE, alpha=0.6, s=20, zorder=3)

    ax1.set_xticks([1, 2])
    ax1.set_xticklabels([f"Non-diminutive\n(n={results['n_nondim']})",
                         f"Diminutive\n(n={results['n_dim']})"])
    ax1.set_ylabel("Top-match score")
    ax1.set_title("Score Distribution by Group", fontweight="bold")
    ax1.set_ylim(-0.05, 1.1)
    ax1.spines[["top", "right"]].set_visible(False)

    # ── 2. Histogram overlay ──────────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.set_facecolor("#F8F9FA")

    bins = np.linspace(0, 1, 25)
    ax2.hist(nondim_scores, bins=bins, alpha=0.6, color=BLUE,
             label=f"Non-diminutive (n={results['n_nondim']})", density=True)
    ax2.hist(dim_scores, bins=bins, alpha=0.7, color=ORANGE,
             label=f"Diminutive (n={results['n_dim']})", density=True)
    ax2.axvline(results["nondim_mean"], color=BLUE,   linestyle="--", linewidth=1.5)
    ax2.axvline(results["dim_mean"],    color=ORANGE, linestyle="--", linewidth=1.5)
    ax2.set_xlabel("Top-match score")
    ax2.set_ylabel("Density")
    ax2.set_title("Score Distributions (Density)", fontweight="bold")
    ax2.legend(fontsize=9)
    ax2.spines[["top", "right"]].set_visible(False)

    # ── 3. Stats summary box ──────────────────────────────────────────────────
    ax3 = fig.add_subplot(gs[1, 0])
    ax3.set_facecolor("#F8F9FA")
    ax3.axis("off")

    p = results["p_value"]
    sig = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "n.s."
    effect = ("|r| < 0.1: negligible" if abs(results["effect_size_r"]) < 0.1
              else "|r| < 0.3: small" if abs(results["effect_size_r"]) < 0.3
              else "|r| < 0.5: medium" if abs(results["effect_size_r"]) < 0.5
              else "|r| ≥ 0.5: large")

    summary = (
        f"Mann-Whitney U Test\n"
        f"{'─' * 36}\n"
        f"  Non-diminutive   mean = {results['nondim_mean']:.4f}\n"
        f"                 median = {results['nondim_median']:.4f}\n"
        f"                    std = {results['nondim_std']:.4f}\n\n"
        f"  Diminutive       mean = {results['dim_mean']:.4f}\n"
        f"                 median = {results['dim_median']:.4f}\n"
        f"                    std = {results['dim_std']:.4f}\n\n"
        f"  U statistic          = {results['u_stat']:.1f}\n"
        f"  p-value              = {p:.4f}  {sig}\n"
        f"  Effect size (r)      = {results['effect_size_r']:.4f}\n"
        f"  Interpretation       : {effect}\n\n"
        f"  * p<0.05  ** p<0.01  *** p<0.001  n.s. = not significant"
    )

    ax3.text(0.05, 0.95, summary, transform=ax3.transAxes,
             fontsize=10, verticalalignment="top", fontfamily="monospace",
             bbox=dict(boxstyle="round,pad=0.6", facecolor="white",
                       edgecolor="#CCCCCC", linewidth=1.5))

    # ── 4. Diminutive word scores (dot plot) ──────────────────────────────────
    ax4 = fig.add_subplot(gs[1, 1])
    ax4.set_facecolor("#F8F9FA")

    dim_sorted = sorted(dim, key=lambda x: x[1])
    words  = [w for w, _ in dim_sorted]
    scores = [s for _, s in dim_sorted]

    ax4.barh(range(len(words)), scores, color=ORANGE, alpha=0.75, height=0.6)
    ax4.axvline(results["nondim_mean"], color=BLUE, linestyle="--",
                linewidth=1.5, label=f"Non-dim mean ({results['nondim_mean']:.3f})")
    ax4.set_yticks(range(len(words)))
    ax4.set_yticklabels(words, fontsize=8)
    ax4.set_xlabel("Top-match score")
    ax4.set_title("Individual Diminutive Word Scores", fontweight="bold")
    ax4.legend(fontsize=9)
    ax4.set_xlim(0, 1.05)
    ax4.spines[["top", "right"]].set_visible(False)

    # ── title ─────────────────────────────────────────────────────────────────
    fig.suptitle("Diminutive vs Non-Diminutive: Top-Match Cosine Similarity",
                 fontsize=13, fontweight="bold", y=1.01)

    plt.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    print(f"  Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  default="results/top10_english_for_polish.xlsx")
    parser.add_argument("--output", default="results/diminutive_analysis.png")
    args = parser.parse_args()

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)

    print(f"Loading {args.input}...")
    dim, nondim = load_data(args.input)
    print(f"  Diminutive words:     {len(dim)}")
    print(f"  Non-diminutive words: {len(nondim)}")

    dim_scores    = [s for _, s in dim]
    nondim_scores = [s for _, s in nondim]

    print("\nRunning Mann-Whitney U test...")
    results = run_stats(dim_scores, nondim_scores)

    p = results["p_value"]
    sig = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "n.s."
    print(f"  U = {results['u_stat']:.1f}, p = {p:.4f} {sig}")
    print(f"  Effect size r = {results['effect_size_r']:.4f}")
    print(f"  Diminutive mean:     {results['dim_mean']:.4f}")
    print(f"  Non-diminutive mean: {results['nondim_mean']:.4f}")

    print("\nGenerating plot...")
    plot(dim, nondim, results, args.output)
    print("\nDone.")


if __name__ == "__main__":
    main()
