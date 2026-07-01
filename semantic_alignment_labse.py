"""
semantic_alignment_labse.py
============================
Combined approach:
  - Paper 2 (Hämmerl et al., 2022): Multilingual contextual model (LaBSE) 
    with subword tokenization for complete CDI word coverage
  - Paper 1 (Liu et al., 2025): Similarity profiles across shared associates 
    + Pearson R_c for cross-lingual word similarity scoring

Dictionary is cleaned at startup: any CDI words found in the dictionary 
are removed so words don't appear in their own similarity profiles.

Outputs:
  data/labse_alignment_scores.csv
  results/top10_english_for_polish_labse.xlsx
  results/top10_polish_for_english_labse.xlsx
  data/dictionary_coverage.txt
  data/missing_from_dictionary_pl.txt
  data/missing_from_dictionary_en.txt

Usage:
  python semantic_alignment_labse.py --top 10
"""

import argparse
import csv
import os
import re
import sys
from collections import defaultdict

import numpy as np
from scipy.stats import pearsonr
from sentence_transformers import SentenceTransformer

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────

BASE_DIR = r"C:\Users\CAyre\Documents\Coding\TranslationClassification\TranslationClassification"
DATA_DIR = os.path.join(BASE_DIR, "data")
RESULTS_DIR = os.path.join(BASE_DIR, "results")

POLISH_CDI = os.path.join(DATA_DIR, "cdi_pl_diminutives.csv")
ENGLISH_CDI = os.path.join(DATA_DIR, "american_english_itemdata.csv")
DICTIONARY = os.path.join(DATA_DIR, "pl_en_dict_clean.csv")

OUTPUT_CSV = os.path.join(DATA_DIR, "labse_alignment_scores.csv")
OUTPUT_EN_FOR_PL = os.path.join(RESULTS_DIR, "top10_english_for_polish_labse.xlsx")
OUTPUT_PL_FOR_EN = os.path.join(RESULTS_DIR, "top10_polish_for_english_labse.xlsx")
DICT_COVERAGE_TXT = os.path.join(DATA_DIR, "dictionary_coverage.txt")
MISSING_PL_TXT = os.path.join(DATA_DIR, "missing_from_dictionary_pl.txt")
MISSING_EN_TXT = os.path.join(DATA_DIR, "missing_from_dictionary_en.txt")

# ── Excel styles ──────────────────────────────────────────────────────────────

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BODY = Font(name="Arial", size=10)


def get_fill(score):
    if score >= 0.75:
        return FILL_GREEN
    elif score >= 0.60:
        return FILL_YELLOW
    else:
        return FILL_RED


# ── helpers ───────────────────────────────────────────────────────────────────

def strip_parens(text):
    return re.sub(r"\s*\(.*?\)", "", text).strip()


def load_cdi(path):
    """Load CDI words from CSV. Expects 'item_definition' and 'category' columns."""
    seen, items = set(), []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            word = strip_parens(row["item_definition"])
            cat = row.get("category", "")
            if word not in seen:
                seen.add(word)
                items.append((word, cat))
    return items


def load_dictionary(path, pl_cdi_words=None, en_cdi_words=None):
    """
    Load the shared associate dictionary.
    Removes any entries where the English or Polish word is a CDI word
    (to prevent words from appearing in their own similarity profiles).
    
    Returns: 
      associates: list of dicts with 'en' and 'pl_list' keys
      en_words_set: set of all English words in the dictionary
      pl_words_set: set of all Polish words in the dictionary
    """
    mapping = defaultdict(list)
    en_words = set()
    pl_words = set()
    
    pl_cdi_lower = set(w.lower() for w in (pl_cdi_words or []))
    en_cdi_lower = set(w.lower() for w in (en_cdi_words or []))
    
    removed_en = set()
    removed_pl = set()
    
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            en_word = strip_parens(row["english"]).lower()
            pl_word = strip_parens(row["polish"]).lower()
            
            # Collect alternatives
            alt_words = []
            if row.get("alternatives"):
                for alt in row["alternatives"].split("|"):
                    alt = strip_parens(alt).lower()
                    if alt and alt != pl_word and alt not in pl_cdi_lower:
                        alt_words.append(alt)
                    elif alt in pl_cdi_lower:
                        removed_pl.add(alt)
            
            # Skip if English word is a CDI word
            if en_word in en_cdi_lower:
                removed_en.add(en_word)
                continue
            
            # Handle Polish word being a CDI word
            if pl_word in pl_cdi_lower:
                removed_pl.add(pl_word)
                if alt_words:
                    pl_word = alt_words[0]
                    alt_words = alt_words[1:]
                else:
                    continue  # No valid Polish translation remains
            
            en_words.add(en_word)
            pl_words.add(pl_word)
            mapping[en_word].append(pl_word)
            
            for alt in alt_words:
                if alt not in pl_cdi_lower:
                    mapping[en_word].append(alt)
                    pl_words.add(alt)

    associates = []
    for en_word, pl_words_list in mapping.items():
        associates.append({"en": en_word, "pl_list": list(set(pl_words_list))})
    
    print(f"  Original dictionary entries loaded")
    if removed_en:
        print(f"  Removed {len(removed_en)} English entries (in CDI)")
    if removed_pl:
        print(f"  Removed {len(removed_pl)} Polish entries (in CDI)")
    print(f"  Final dictionary size: {len(associates)} entries")
    
    return associates, en_words, pl_words


# ── LaBSE embedding ──────────────────────────────────────────────────────────

def embed_words_labse(words):
    """
    Embed words using LaBSE with template sentences (Paper 2 insight).
    Returns dict: {word: normalized_vector}
    """
    print("  Loading LaBSE model...")
    model = SentenceTransformer("sentence-transformers/LaBSE")
    print("  Model ready.\n")

    templates = [
        "This is a {word}.",
        "I see a {word}.",
        "Where is the {word}?",
    ]

    words_list = list(words)
    all_sentences = []
    word_to_indices = defaultdict(list)

    for word in words_list:
        for t in templates:
            word_to_indices[word].append(len(all_sentences))
            all_sentences.append(t.format(word=word))

    print(f"  Embedding {len(all_sentences)} sentences...")
    embeddings = model.encode(
        all_sentences,
        normalize_embeddings=True,
        show_progress_bar=True,
        convert_to_numpy=True
    )

    word_vecs = {}
    for word in words_list:
        indices = word_to_indices[word]
        avg_vec = np.mean(embeddings[indices], axis=0)
        avg_vec = avg_vec / (np.linalg.norm(avg_vec) + 1e-10)
        word_vecs[word] = avg_vec

    print(f"  Embedded {len(word_vecs)}/{len(words_list)} words")
    return word_vecs


# ── core algorithm (Paper 1) ─────────────────────────────────────────────────

def build_associate_vectors(associates, pl_vecs, en_vecs):
    """
    For each shared associate, get its English vector and the average
    of its Polish equivalent vectors (Formula 1 from Liu et al., 2025).
    """
    en_vec_list = []
    pl_vec_list = []

    for assoc in associates:
        en_word = assoc["en"].lower()
        pl_words = [w.lower() for w in assoc["pl_list"]]

        en_vec = en_vecs.get(en_word)
        if en_vec is None:
            continue

        pl_vecs_for_word = []
        for pw in pl_words:
            if pw in pl_vecs:
                pl_vecs_for_word.append(pl_vecs[pw])

        if not pl_vecs_for_word:
            continue

        pl_vec = np.mean(pl_vecs_for_word, axis=0)
        pl_vec = pl_vec / (np.linalg.norm(pl_vec) + 1e-10)

        en_vec_list.append(en_vec)
        pl_vec_list.append(pl_vec)

    print(f"  Valid shared associates: {len(en_vec_list)}/{len(associates)}")
    return np.array(en_vec_list), np.array(pl_vec_list)


def semantic_alignment_score(pl_word, en_word, pl_vecs, en_vecs,
                              en_assoc_vecs, pl_assoc_vecs):
    """Calculate Pearson R_c between two words' similarity profiles."""
    pl_vec = pl_vecs.get(pl_word.lower())
    en_vec = en_vecs.get(en_word.lower())

    if pl_vec is None or en_vec is None:
        return None

    sims_pl = pl_assoc_vecs @ pl_vec
    sims_en = en_assoc_vecs @ en_vec

    if len(sims_pl) < 3:
        return None

    r_c, _ = pearsonr(sims_en, sims_pl)
    return r_c


# ── dictionary coverage analysis ─────────────────────────────────────────────

def analyze_dictionary_coverage(cdi_words, dict_words, language_name, output_path):
    """
    Check which CDI words are in the dictionary and which aren't.
    Saves report to output_path.
    """
    cdi_set = set(w.lower() for w in cdi_words)
    dict_set = set(w.lower() for w in dict_words)
    
    in_dict = cdi_set & dict_set
    missing = cdi_set - dict_set
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"Dictionary Coverage for {language_name} CDI Words\n")
        f.write(f"{'=' * 50}\n\n")
        f.write(f"Total CDI words: {len(cdi_set)}\n")
        f.write(f"Words in dictionary: {len(in_dict)} ({100*len(in_dict)/len(cdi_set):.1f}%)\n")
        f.write(f"Words MISSING from dictionary: {len(missing)} ({100*len(missing)/len(cdi_set):.1f}%)\n\n")
        
        f.write(f"{'─' * 50}\n")
        f.write(f"MISSING WORDS:\n")
        f.write(f"{'─' * 50}\n")
        for w in sorted(missing):
            f.write(f"  {w}\n")
    
    print(f"\n  ── {language_name} Dictionary Coverage ──")
    print(f"  CDI words: {len(cdi_set)}")
    print(f"  In dictionary: {len(in_dict)} ({100*len(in_dict)/len(cdi_set):.1f}%)")
    print(f"  Missing: {len(missing)} ({100*len(missing)/len(cdi_set):.1f}%)")
    print(f"  Full list saved to: {output_path}")
    
    return in_dict, missing


# ── main pipeline ─────────────────────────────────────────────────────────────

def run_pipeline(top_n=10):
    # 1. Load data
    print("=" * 60)
    print("STEP 1: Loading CDI data and dictionary")
    print("=" * 60)

    pl_items = load_cdi(POLISH_CDI)
    en_items = load_cdi(ENGLISH_CDI)

    pl_words = [w for w, _ in pl_items]
    en_words = [w for w, _ in en_items]

    print(f"  Polish CDI words: {len(pl_words)}")
    print(f"  English CDI words: {len(en_words)}")

    # Load dictionary, removing CDI words from it
    associates, dict_en_words, dict_pl_words = load_dictionary(
        DICTIONARY,
        pl_cdi_words=pl_words,
        en_cdi_words=en_words
    )

    print(f"  Total pairs to score: {len(pl_words) * len(en_words):,}")

    # ── Dictionary coverage analysis ──────────────────────────────────────
    print("\n" + "=" * 60)
    print("STEP 1B: Dictionary coverage analysis")
    print("=" * 60)

    pl_in_dict, pl_missing_dict = analyze_dictionary_coverage(
        pl_words, dict_pl_words, "Polish", MISSING_PL_TXT
    )
    en_in_dict, en_missing_dict = analyze_dictionary_coverage(
        en_words, dict_en_words, "English", MISSING_EN_TXT
    )

    # Combined report
    with open(DICT_COVERAGE_TXT, "w", encoding="utf-8") as f:
        f.write("DICTIONARY COVERAGE SUMMARY\n")
        f.write(f"{'=' * 50}\n\n")
        f.write(f"Dictionary entries (after removing CDI words): {len(associates)}\n\n")
        f.write(f"Polish CDI words: {len(pl_words)}\n")
        f.write(f"  In dictionary: {len(pl_in_dict)} ({100*len(pl_in_dict)/len(pl_words):.1f}%)\n")
        f.write(f"  Missing: {len(pl_missing_dict)} ({100*len(pl_missing_dict)/len(pl_words):.1f}%)\n\n")
        f.write(f"English CDI words: {len(en_words)}\n")
        f.write(f"  In dictionary: {len(en_in_dict)} ({100*len(en_in_dict)/len(en_words):.1f}%)\n")
        f.write(f"  Missing: {len(en_missing_dict)} ({100*len(en_missing_dict)/len(en_words):.1f}%)\n")

    print(f"\n  Coverage summary saved to: {DICT_COVERAGE_TXT}")

    # 2. Embed with LaBSE
    print("\n" + "=" * 60)
    print("STEP 2: Embedding words with LaBSE")
    print("=" * 60)

    dict_pl_list = []
    for a in associates:
        dict_pl_list.extend(a["pl_list"])
    dict_en_list = [a["en"] for a in associates]

    all_pl = list(set(pl_words + dict_pl_list))
    all_en = list(set(en_words + dict_en_list))

    print(f"  Polish words to embed: {len(all_pl)}")
    print(f"  English words to embed: {len(all_en)}\n")

    all_words = list(set(all_pl + all_en))
    all_vecs = embed_words_labse(all_words)

    pl_vecs = {w.lower(): all_vecs[w] for w in all_pl if w in all_vecs}
    en_vecs = {w.lower(): all_vecs[w] for w in all_en if w in all_vecs}

    # Report embedding coverage
    pl_missing_emb = [w for w in pl_words if w.lower() not in pl_vecs]
    en_missing_emb = [w for w in en_words if w.lower() not in en_vecs]

    if pl_missing_emb:
        print(f"\n  ⚠ Polish CDI words not embedded: {len(pl_missing_emb)}")
    else:
        print(f"\n  ✓ All {len(pl_words)} Polish CDI words embedded")

    if en_missing_emb:
        print(f"  ⚠ English CDI words not embedded: {len(en_missing_emb)}")
    else:
        print(f"  ✓ All {len(en_words)} English CDI words embedded")

    # 3. Build associate vectors (once, with cleaned dictionary)
    print("\n" + "=" * 60)
    print("STEP 3: Building shared associate vectors")
    print("=" * 60)

    en_assoc_vecs, pl_assoc_vecs = build_associate_vectors(
        associates, pl_vecs, en_vecs
    )

    if len(en_assoc_vecs) < 3:
        print("ERROR: Too few valid shared associates.")
        return

    # 4. Compute semantic alignment scores
    print("\n" + "=" * 60)
    print("STEP 4: Computing semantic alignment scores (Pearson R_c)")
    print("=" * 60)

    rows = []
    total = len(pl_words) * len(en_words)
    count = 0

    pl_lower_in_dict = set(w.lower() for w in pl_in_dict)
    en_lower_in_dict = set(w.lower() for w in en_in_dict)

    for pl_w, pl_cat in pl_items:
        for en_w, en_cat in en_items:
            r_c = semantic_alignment_score(
                pl_w, en_w, pl_vecs, en_vecs,
                en_assoc_vecs, pl_assoc_vecs
            )
            if r_c is not None and not np.isnan(r_c):
                rows.append({
                    "polish_word": pl_w,
                    "english_word": en_w,
                    "polish_category": pl_cat,
                    "english_category": en_cat,
                    "R_c": round(r_c, 6),
                    "pl_in_dict": pl_w.lower() in pl_lower_in_dict,
                    "en_in_dict": en_w.lower() in en_lower_in_dict,
                })
            count += 1
            if count % 10000 == 0:
                print(f"  Processed {count}/{total} pairs...")

    print(f"  Scored: {len(rows)}, Skipped: {total - len(rows)}")

    # 5. Save CSV
    print(f"\nWriting {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "polish_word", "english_word", "polish_category",
            "english_category", "R_c", "pl_in_dict", "en_in_dict"
        ])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    # 6. Excel output
    print("\n" + "=" * 60)
    print("STEP 5: Building Excel output")
    print("=" * 60)

    pl_groups = defaultdict(list)
    en_groups = defaultdict(list)

    for r in rows:
        pl_groups[r["polish_word"]].append({
            "match_word": r["english_word"],
            "match_cat": r["english_category"],
            "index_cat": r["polish_category"],
            "score": r["R_c"],
            "index_in_dict": r["pl_in_dict"],
            "match_in_dict": r["en_in_dict"],
        })
        en_groups[r["english_word"]].append({
            "match_word": r["polish_word"],
            "match_cat": r["polish_category"],
            "index_cat": r["english_category"],
            "score": r["R_c"],
            "index_in_dict": r["en_in_dict"],
            "match_in_dict": r["pl_in_dict"],
        })

    write_excel(pl_groups, top_n, OUTPUT_EN_FOR_PL,
                "polish_word", "polish_category",
                "english_word", "english_category")

    write_excel(en_groups, top_n, OUTPUT_PL_FOR_EN,
                "english_word", "english_category",
                "polish_word", "polish_category")

    # 7. Print top pairs
    rows_sorted = sorted(rows, key=lambda r: r["R_c"], reverse=True)
    print(f"\nTop 25 most semantically aligned pairs (LaBSE + Paper 1):")
    print(f"  {'Polish':<30} {'English':<30} {'R_c':>10}  in_dict")
    print("  " + "-" * 80)
    for r in rows_sorted[:25]:
        dict_flag = f"pl={'Y' if r['pl_in_dict'] else 'N'} en={'Y' if r['en_in_dict'] else 'N'}"
        print(f"  {r['polish_word']:<30} {r['english_word']:<30} {r['R_c']:>10.4f}  {dict_flag}")

    print(f"\nAll done!")
    print(f"  Scores: {OUTPUT_CSV}")
    print(f"  Excel:  {OUTPUT_EN_FOR_PL}")
    print(f"  Excel:  {OUTPUT_PL_FOR_EN}")
    print(f"  Coverage: {DICT_COVERAGE_TXT}")
    print(f"  Missing PL: {MISSING_PL_TXT}")
    print(f"  Missing EN: {MISSING_EN_TXT}")


def write_excel(groups, top_n, output_path, index_col, index_cat_col,
                match_col, match_cat_col):
    wb = Workbook()
    ws = wb.active
    ws.title = "Top matches"

    headers = [index_col, index_cat_col, "rank", match_col, match_cat_col,
               "R_c", "index_in_dict", "match_in_dict"]
    col_widths = [30, 22, 6, 30, 22, 12, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    row_num = 2
    for index_word, matches in sorted(groups.items()):
        top = sorted(matches, key=lambda x: x["score"], reverse=True)[:top_n]
        for rank, m in enumerate(top, 1):
            values = [
                index_word, m["index_cat"], rank, m["match_word"],
                m["match_cat"], round(m["score"], 6),
                m.get("index_in_dict", "N/A"),
                m.get("match_in_dict", "N/A"),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = get_fill(m["score"])
                cell.font = FONT_BODY
                cell.alignment = Alignment(
                    horizontal="center" if col_idx in (3, 7, 8) else "left"
                )
            row_num += 1

    wb.save(output_path)
    print(f"  Written: {output_path}  ({row_num - 2} rows)")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Semantic Alignment with LaBSE (Papers 1 + 2 combined)"
    )
    parser.add_argument("--top", type=int, default=10,
                        help="Number of top matches per word (default 10)")
    args = parser.parse_args()

    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    for path, name in [
        (POLISH_CDI, "Polish CDI"),
        (ENGLISH_CDI, "English CDI"),
        (DICTIONARY, "Dictionary"),
    ]:
        if not os.path.exists(path):
            print(f"ERROR: {name} not found: {path}")
            sys.exit(1)

    print("Semantic Alignment Pipeline (LaBSE + Paper 1)")
    print("=" * 50)
    print()
    print("Output files:")
    print(f"  Scores:  {OUTPUT_CSV}")
    print(f"  Excel:   {OUTPUT_EN_FOR_PL}")
    print(f"  Excel:   {OUTPUT_PL_FOR_EN}")
    print(f"  Coverage: {DICT_COVERAGE_TXT}")
    print()

    run_pipeline(top_n=args.top)


if __name__ == "__main__":
    main()